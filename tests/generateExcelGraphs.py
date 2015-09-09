import openpyxl
import os
import sys

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import pytweening

MAKE_CHARTS = False # seems to be a bug in openpyxl, it produces corrupt spreadsheet files.

wb = openpyxl.Workbook()
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

graphs = (('Quad', pytweening.easeInQuad, pytweening.easeOutQuad, pytweening.easeInOutQuad),
          ('Cubic', pytweening.easeInCubic, pytweening.easeOutCubic, pytweening.easeInOutCubic),
          ('Quart', pytweening.easeInQuart, pytweening.easeOutQuart, pytweening.easeInOutQuart),
          ('Quint', pytweening.easeInQuint, pytweening.easeOutQuint, pytweening.easeInOutQuint),
          ('Sine', pytweening.easeInSine, pytweening.easeOutSine, pytweening.easeInOutSine),
          ('Expo', pytweening.easeInExpo, pytweening.easeOutExpo, pytweening.easeInOutExpo),
          ('Circ', pytweening.easeInCirc, pytweening.easeOutCirc, pytweening.easeInOutCirc),
          ('Elastic', pytweening.easeInElastic, pytweening.easeOutElastic, pytweening.easeInOutElastic),
          ('Back', pytweening.easeInBack, pytweening.easeOutBack, pytweening.easeInOutBack),
          ('Bounce', pytweening.easeInBounce, pytweening.easeOutBounce, pytweening.easeInOutBounce),
         )

# Linear function has only a single function, not three.
wb.create_sheet(title='Linear')
sheet = wb.get_sheet_by_name('Linear')
for i in range(1, 101):
    n = i / 100.0
    sheet['A' + str(i)] = pytweening.linear(n)

    if MAKE_CHARTS:
        refObj = openpyxl.charts.Reference(sheet, (1, 1), (100, 1))
        seriesObj = openpyxl.charts.Series(refObj, title='Linear')
        chartObj = openpyxl.charts.LineChart()
        chartObj.append(seriesObj)
        chartObj.drawing.top = 50
        chartObj.drawing.left = 300
        chartObj.drawing.width = 300
        chartObj.drawing.height = 200
        sheet.add_chart(chartObj)


for graph in graphs:
    name, easeInFunc, easeOutFunc, easeInOutFunc = graph
    wb.create_sheet(title=name)
    sheet = wb.get_sheet_by_name(name)
    for i in range(1, 101):
        n = i / 100.0
        sheet['A' + str(i)] = easeInFunc(n)
        sheet['B' + str(i)] = easeOutFunc(n)
        sheet['C' + str(i)] = easeInOutFunc(n)

        if MAKE_CHARTS:
            inRefObj = openpyxl.charts.Reference(sheet, (1, 1), (100, 1))
            inSeriesObj = openpyxl.charts.Series(inRefObj, title=name + ' In')
            outRefObj = openpyxl.charts.Reference(sheet, (1, 2), (100, 2))
            outSeriesObj = openpyxl.charts.Series(outRefObj, title=name + ' Out')
            intOutRefObj = openpyxl.charts.Reference(sheet, (1, 3), (100, 3))
            inOutSeriesObj = openpyxl.charts.Series(intOutRefObj, title=name + ' InOut')

            chartObj = openpyxl.charts.LineChart()
            chartObj.append(inSeriesObj)
            chartObj.append(outSeriesObj)
            chartObj.append(inOutSeriesObj)
            chartObj.drawing.top = 50
            chartObj.drawing.left = 300
            chartObj.drawing.width = 300
            chartObj.drawing.height = 200
            sheet.add_chart(chartObj)


wb.save('graphs.xlsx')
